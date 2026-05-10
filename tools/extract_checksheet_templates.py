import json
import re
import sys
from pathlib import Path

import openpyxl


PHOTO_LIMIT = 4

ID_OVERRIDES = {
    "C1 & C2 Civil Check Sheet-A R1.xlsx": "tpl-c2-civil-stage-audit-r1",
}

COMMON_FIELDS = [
    {"id": "site_id", "label": "Site ID", "type": "text", "required": True},
    {"id": "site_name", "label": "Site Name", "type": "text"},
    {"id": "country_circle", "label": "Country / Circle", "type": "text"},
    {"id": "district_area", "label": "District / Area", "type": "text"},
    {"id": "site_type", "label": "Site Type", "type": "text"},
    {"id": "tower_type", "label": "Tower Type", "type": "text"},
    {"id": "solution_type", "label": "Solution Type", "type": "text"},
    {"id": "build_partner", "label": "Build Partner", "type": "text"},
    {"id": "quality_audit_agency", "label": "Quality Audit Agency", "type": "text"},
    {"id": "audit_engineer", "label": "Audit Engineer", "type": "text", "required": True},
    {"id": "audit_date", "label": "Date of Audit", "type": "date", "required": True},
    {"id": "audit_start_time", "label": "Audit Start Time", "type": "time"},
    {"id": "audit_end_time", "label": "Audit End Time", "type": "time"},
]

SKIP_FIELD_TERMS = {
    "indus towers ltd",
    "indus tower ltd",
    "indus towers ltd.",
    "indus tower ltd.",
    "indus towers ltd",
    "indus towers ltd.",
    "infra quality technology",
    "photo along with date, time and lat long stamped",
}


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ")
    text = re.sub(r"_+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_id(value):
    text = clean_text(value)
    if re.match(r"^\d+\.0$", text):
        return text[:-2]
    return text


def slugify(value):
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text[:72] or "item"


def unique_slug(base, seen):
    slug = slugify(base)
    candidate = slug
    index = 2
    while candidate in seen:
        candidate = f"{slug}-{index}"
        index += 1
    seen.add(candidate)
    return candidate


def row_values(ws, row_index):
    return [clean_text(ws.cell(row_index, col).value) for col in range(1, ws.max_column + 1)]


def meaningful_cells(row):
    return [cell for cell in row if cell]


def first_meaningful_title(workbook, path):
    fallback = path.stem
    candidates = []
    for ws in workbook.worksheets:
        for row_index in range(1, min(ws.max_row, 8) + 1):
            cells = meaningful_cells(row_values(ws, row_index))
            if not cells:
                continue
            text = cells[0]
            low = text.lower()
            if any(term in low for term in ["check", "acceptance", "audit", "inventory", "report"]):
                candidates.append(text)
    return candidates[0] if candidates else fallback


def add_field(fields, label, field_type="text", required=False):
    label = clean_text(label).strip(" :-")
    if not label:
        return
    low = label.lower()
    if low in SKIP_FIELD_TERMS:
        return
    if any(term in low for term in ["photograph", "photo placeholder", "photo along"]):
        return
    if len(label) > 70:
        return
    field_id = slugify(label)
    if any(field["id"] == field_id or field["label"].lower() == low for field in fields):
        return
    fields.append({"id": field_id, "label": label, "type": field_type, "required": required})


def extract_field_labels_from_cell(cell):
    text = clean_text(cell)
    if not text:
        return []
    labels = []
    colon_matches = re.findall(r"([A-Za-z0-9/&(). -]{2,45})\s*:", text)
    if colon_matches:
        for match in colon_matches:
            label = clean_text(match).strip(" :-")
            if label:
                labels.append(label)
        return labels
    if len(text) <= 35 and re.search(r"\b(country|area|district)\b", text, re.I):
        return [text]
    return []


def first_content_row(workbook):
    first = None
    for ws in workbook.worksheets:
        for row_index in range(1, ws.max_row + 1):
            joined = " ".join(meaningful_cells(row_values(ws, row_index))).lower()
            if not joined:
                continue
            if "photograph" in joined or "photo placeholder" in joined:
                first = row_index if first is None else min(first, row_index)
                break
            row = [cell.lower() for cell in row_values(ws, row_index)]
            if is_header_row(row) or is_section_title(row[0] if row else ""):
                first = row_index if first is None else min(first, row_index)
                break
    return first or 12


def extract_site_fields(workbook):
    fields = [dict(field) for field in COMMON_FIELDS]
    for ws in workbook.worksheets:
        stop = min(ws.max_row, first_content_row(workbook) + 2, 18)
        for row_index in range(1, stop + 1):
            for cell in meaningful_cells(row_values(ws, row_index)):
                for label in extract_field_labels_from_cell(cell):
                    field_type = "date" if "date" in label.lower() else "time" if "time" in label.lower() else "text"
                    add_field(fields, label, field_type=field_type)
    return fields


def is_section_title(value):
    return bool(re.match(r"^\d+\.\s+\S", clean_text(value)))


def is_header_row(row):
    low = [clean_text(cell).lower() for cell in row]
    has_id = any(
        cell in {"id", "s.no", "s.no.", "sr no", "sl no", "ref"} or cell.startswith("ref")
        for cell in low
    )
    has_item = any(
        any(term in cell for term in ["checkpoint", "description", "requirement", "item", "parameter"])
        for cell in low
    )
    return has_id and has_item and not any("photo description" in cell for cell in low)


def find_col(row, terms, fallback=None):
    low = [clean_text(cell).lower() for cell in row]
    for term in terms:
        for index, cell in enumerate(low):
            if term in cell:
                return index
    return fallback


def nearest_section(ws, header_row):
    for row_index in range(header_row - 1, max(1, header_row - 10), -1):
        cells = meaningful_cells(row_values(ws, row_index))
        if not cells:
            continue
        first = cells[0]
        if is_section_title(first):
            description = ""
            if row_index + 1 < header_row:
                desc_cells = meaningful_cells(row_values(ws, row_index + 1))
                if desc_cells and not is_header_row(desc_cells):
                    description = desc_cells[0]
            return first, description
    for row_index in range(1, min(ws.max_row, 6) + 1):
        cells = meaningful_cells(row_values(ws, row_index))
        if cells and any(term in cells[0].lower() for term in ["check", "acceptance", "audit", "inventory"]):
            return cells[0], ""
    return ws.title, ""


def is_stop_row(row):
    joined = " ".join(meaningful_cells(row)).lower()
    first = row[0] if row else ""
    if not joined:
        return False
    return (
        is_section_title(first)
        or is_header_row(row)
        or "photograph" in joined
        or "photo placeholder" in joined
        or "signature" in joined
        or "audit representative" in joined
    )


def extract_structured_sections(workbook):
    sections = []
    section_ids = set()
    for ws in workbook.worksheets:
        for row_index in range(1, ws.max_row + 1):
            header = row_values(ws, row_index)
            if not is_header_row(header):
                continue
            low_header = [cell.lower() for cell in header]
            if any("photo description" in cell for cell in low_header):
                continue
            id_col = find_col(header, ["s.no", "sr no", "sl no", "id", "ref"], 0)
            item_col = find_col(header, ["checkpoint", "description", "item", "parameter"], 1)
            criteria_col = find_col(header, ["requirement", "criteria"], None)
            input_headers = [
                clean_text(cell)
                for cell in header
                if clean_text(cell)
                and not any(term in clean_text(cell).lower() for term in ["id", "s.no", "checkpoint", "description", "requirement", "criteria", "status", "compliance"])
            ]
            input_label = " / ".join(input_headers) if input_headers else "Remarks"

            title, description = nearest_section(ws, row_index)
            if "inventory" in title.lower() or "inventory" in workbook.sheetnames[0].lower():
                title = "Inventory Items"
                description = "Record make, serial number, size/capacity, quantity, and remarks for each site item."
                input_label = "Make / Sr. No / Size / Quantity / Remarks"

            items = []
            cursor = row_index + 1
            while cursor <= ws.max_row:
                row = row_values(ws, cursor)
                if is_stop_row(row):
                    break
                cells = meaningful_cells(row)
                if cells:
                    item_id = clean_id(row[id_col] if id_col < len(row) else "")
                    item = clean_text(row[item_col] if item_col is not None and item_col < len(row) else "")
                    if not item and len(cells) > 1:
                        item = cells[1]
                    criteria = clean_text(row[criteria_col] if criteria_col is not None and criteria_col < len(row) else "")
                    if item:
                        if not criteria:
                            other_values = []
                            for idx, cell in enumerate(row):
                                if idx not in {id_col, item_col} and clean_text(cell):
                                    other_values.append(f"{header[idx] or 'Value'}: {cell}")
                            criteria = "; ".join(other_values) or "Fill the required field values and remarks."
                        items.append(
                            {
                                "id": item_id or f"{len(items) + 1}",
                                "item": item,
                                "criteria": criteria,
                                "inputLabel": input_label,
                            }
                        )
                cursor += 1

            if items:
                section_id = unique_slug(title, section_ids)
                sections.append(
                    {
                        "id": section_id,
                        "title": title,
                        "description": description,
                        "items": items,
                    }
                )
    return sections


def extract_installation_requirement_sections(workbook, existing_sections):
    sections = []
    for ws in workbook.worksheets:
        for row_index in range(1, ws.max_row + 1):
            cells = meaningful_cells(row_values(ws, row_index))
            if not cells or "significant installation phases" not in cells[0].lower():
                continue
            items = []
            cursor = row_index + 1
            while cursor <= ws.max_row:
                row = row_values(ws, cursor)
                joined = " ".join(meaningful_cells(row)).lower()
                if not joined:
                    cursor += 1
                    continue
                if "photograph" in joined or "photo along" in joined or is_section_title(row[0]):
                    break
                for cell in meaningful_cells(row):
                    if len(cell) > 5:
                        label = cell.split(":")[0].strip() if ":" in cell else cell
                        items.append(
                            {
                                "id": f"R-{len(items) + 1:02d}",
                                "item": label,
                                "criteria": cell,
                                "inputLabel": "Value / Remarks",
                            }
                        )
                cursor += 1
            if items:
                sections.append(
                    {
                        "id": unique_slug(f"{ws.title} installation requirements", {s["id"] for s in existing_sections + sections}),
                        "title": "Significant Installation Phases & Requirements",
                        "description": "Capture the listed installation values or remarks from the original check sheet.",
                        "items": items,
                    }
                )
    return sections


def parse_placeholder_photos(workbook, seen_keys):
    photos = []
    used_ids = set()
    for ws in workbook.worksheets:
        row_index = 1
        while row_index <= ws.max_row:
            row = row_values(ws, row_index)
            cells = meaningful_cells(row)
            placeholder = next((cell for cell in cells if re.search(r"photo placeholder", cell, re.I)), "")
            if not placeholder:
                row_index += 1
                continue
            title = re.sub(r"^\[?\s*Photo Placeholder\s*\d*\s*:?\s*", "", placeholder, flags=re.I)
            title = title.rstrip(" ]") or placeholder.strip("[]")
            requirements = []
            cursor = row_index + 1
            while cursor <= ws.max_row:
                next_cells = meaningful_cells(row_values(ws, cursor))
                joined = " ".join(next_cells)
                if not joined:
                    cursor += 1
                    continue
                if re.search(r"photo placeholder", joined, re.I) or is_section_title(next_cells[0] if next_cells else ""):
                    break
                if re.search(r"requirement|photograph", joined, re.I):
                    requirements.append(re.sub(r"^Requirement:\s*", "", joined, flags=re.I))
                cursor += 1
            requirement = " ".join(requirements).strip() or placeholder
            key = slugify(title + requirement)
            if key not in seen_keys:
                seen_keys.add(key)
                photos.append({"id": unique_slug(f"photo-{title}", used_ids), "title": title, "requirement": requirement, "limit": PHOTO_LIMIT})
            row_index = max(cursor, row_index + 1)
    return photos


def parse_numbered_photos(workbook, seen_keys):
    photos = []
    used_ids = set()
    pattern = re.compile(r"photograph\s*[- ]*\s*(\d+)", re.I)
    for ws in workbook.worksheets:
        for row_index in range(1, ws.max_row + 1):
            for cell in meaningful_cells(row_values(ws, row_index)):
                if not pattern.search(cell):
                    continue
                if "photo along with date" in cell.lower():
                    continue
                for part in re.split(r"\s{2,}|\s+\|\s+", cell):
                    part = clean_text(part)
                    if not pattern.search(part):
                        continue
                    match = pattern.search(part)
                    number = match.group(1)
                    description = re.sub(r"^photograph\s*[- ]*\s*\d+\s*[-:]?\s*", "", part, flags=re.I).strip()
                    description = description.strip("() ") or f"Photograph {number}"
                    title = f"Photograph {number} - {description}"
                    key = slugify(part)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    photo_id = unique_slug(f"photo-{number}-{description}", used_ids)
                    photos.append({"id": photo_id, "title": title, "requirement": part, "limit": PHOTO_LIMIT})
    return photos


def parse_photo_documentation_list(workbook, seen_keys):
    photos = []
    used_ids = set()
    for ws in workbook.worksheets:
        in_photo_section = False
        row_index = 1
        while row_index <= ws.max_row:
            row = row_values(ws, row_index)
            cells = meaningful_cells(row)
            first = cells[0] if cells else ""
            joined = " ".join(cells).lower()
            if "photographic documentation" in joined or "photographic evidence" in joined:
                in_photo_section = True
                row_index += 1
                continue
            if in_photo_section and re.search(r"audit conclusion|mandatory audit compliance", joined, re.I):
                break
            match = re.match(r"^(\d+)\.\s+(.+)$", first)
            if in_photo_section and match:
                number = match.group(1)
                title_text = match.group(2).strip()
                cursor = row_index + 1
                requirement = ""
                while cursor <= ws.max_row:
                    next_cells = meaningful_cells(row_values(ws, cursor))
                    next_first = next_cells[0] if next_cells else ""
                    if next_cells:
                        if re.match(r"^\d+\.\s+.+$", next_first) or re.search(r"audit conclusion", " ".join(next_cells), re.I):
                            break
                        requirement = " ".join(next_cells)
                        break
                    cursor += 1
                title = f"{number}. {title_text}"
                key = slugify(title + requirement)
                if key not in seen_keys:
                    seen_keys.add(key)
                    photos.append(
                        {
                            "id": unique_slug(f"photo-{number}-{title_text}", used_ids),
                            "title": title,
                            "requirement": requirement or title,
                            "limit": PHOTO_LIMIT,
                        }
                    )
                row_index = max(cursor, row_index + 1)
                continue
            row_index += 1
    return photos


def parse_ref_photo_table(workbook, seen_keys):
    photos = []
    used_ids = set()
    for ws in workbook.worksheets:
        for row_index in range(1, ws.max_row + 1):
            row = row_values(ws, row_index)
            low = [cell.lower() for cell in row]
            if not ("ref" in low and any("photo description" in cell for cell in low)):
                continue
            ref_col = low.index("ref")
            desc_col = find_col(row, ["photo description"], 1)
            req_col = find_col(row, ["detailed requirement", "requirement"], 2)
            cursor = row_index + 1
            while cursor <= ws.max_row:
                data = row_values(ws, cursor)
                ref = clean_text(data[ref_col] if ref_col < len(data) else "")
                desc = clean_text(data[desc_col] if desc_col is not None and desc_col < len(data) else "")
                req = clean_text(data[req_col] if req_col is not None and req_col < len(data) else "")
                if not ref and not desc and not req:
                    cursor += 1
                    continue
                if not re.match(r"^P-\d+", ref, re.I):
                    break
                title = f"{ref} - {desc or 'Photo Evidence'}"
                requirement = req or desc or title
                key = slugify(title + requirement)
                if key not in seen_keys:
                    seen_keys.add(key)
                    photos.append(
                        {
                            "id": unique_slug(f"photo-{ref}-{desc}", used_ids),
                            "title": title,
                            "requirement": requirement,
                            "limit": PHOTO_LIMIT,
                        }
                    )
                cursor += 1
    return photos


def extract_photo_requirements(workbook):
    seen_keys = set()
    ref_photos = parse_ref_photo_table(workbook, seen_keys)
    if ref_photos:
        return ref_photos
    placeholder_photos = parse_placeholder_photos(workbook, seen_keys)
    if placeholder_photos:
        return placeholder_photos
    documentation_photos = parse_photo_documentation_list(workbook, seen_keys)
    if documentation_photos:
        return documentation_photos
    return parse_numbered_photos(workbook, seen_keys)


def build_template(path):
    workbook = openpyxl.load_workbook(path, data_only=False)
    title = first_meaningful_title(workbook, path)
    sections = extract_structured_sections(workbook)
    sections.extend(extract_installation_requirement_sections(workbook, sections))
    template = {
        "id": ID_OVERRIDES.get(path.name, f"tpl-{slugify(path.stem)}"),
        "name": title,
        "revision": "Embedded",
        "source": path.name,
        "siteFields": extract_site_fields(workbook),
        "sections": sections,
        "photoRequirements": extract_photo_requirements(workbook),
    }
    return template


def main():
    source_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\HP\OneDrive\Desktop\Checksheet")
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("templates.js")
    templates = [build_template(path) for path in sorted(source_dir.glob("*.xlsx"))]
    output = "window.CHECKSHEET_TEMPLATES = "
    output += json.dumps(templates, ensure_ascii=False, indent=2)
    output += ";\n"
    output_path.write_text(output, encoding="utf-8")
    summary = [
        {
            "source": template["source"],
            "name": template["name"],
            "sections": len(template["sections"]),
            "items": sum(len(section["items"]) for section in template["sections"]),
            "photoPoints": len(template["photoRequirements"]),
        }
        for template in templates
    ]
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
